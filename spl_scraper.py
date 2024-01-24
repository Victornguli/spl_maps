import asyncio
import json
import logging
from pathlib import Path
from typing import Union

import aiohttp

from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font
from openpyxl.utils import get_column_letter

from bs4 import BeautifulSoup

logging.basicConfig(level=logging.INFO)
_logger = logging.getLogger(__name__)


REGIONS = {
    '1': 'Riyadh',
    '6': 'Asir',
    '5': 'Eastern',
    '8': 'Hail',
    '10': 'Jazan',
    '3': 'Madinah',
    '2': 'Makkah',
    '9': 'Northern Borders',
    '4': 'Al Qassim',
    '7': 'Tabuk',
    '11': 'Najran',
    '12': 'Al Bahah',
    '13': 'Al Jawf',
}

HEADERS = {
    'Referrer': 'https://maps.splonline.com.sa/',
    'Host': 'maps.splonline.com.sa',
    'Origin': 'https://maps.splonline.com.sa',
    'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:109.0) Gecko/20100101 Firefox/117.0',
    'Content-Type': 'application/json, charset=utf-8',
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'Accept-Encoding': 'gzip, deflate, br',
    'X-Requested-With': 'XMLHttpRequest'
}
REGIONS_HTML = 'regions_list.html'
RESULTS_JSON = 'region_cities_districts.json'
RESULTS_EXCEL = 'region_cities_districts.json'


def get_regions(use_arabic: bool = True) -> dict:
    """
    Extract Saudi regions/Provinces. When use_arabic is set parse from the html snippet else
    return hardcoded English names
    :param use_arabic:
    :return:
    """
    if not use_arabic:
        return REGIONS

    regions = {}
    regions_html = None
    html_path = Path(__file__).resolve().parent / REGIONS_HTML
    if html_path.exists():
        with open(html_path, 'r', encoding='utf-8') as f:
            regions_html = f.read()

    if regions_html:
        soup = BeautifulSoup(regions_html, 'lxml')
        options = soup.find_all('option')
        for option in options:
            try:
                region_id = int(option['id'])
            except ValueError as ex:
                _logger.exception(ex)
            else:
                regions[str(region_id)] = option.get_text().strip()

    return regions


async def get_districts(city_id: str, use_arabic: bool = True) -> dict[str, list]:
    async with aiohttp.ClientSession() as session:
        async with session.post(
            'https://maps.splonline.com.sa/Home/GetDistricts',
            data=json.dumps({'cityId': city_id}),
            headers=HEADERS
        ) as response:
            res = json.loads(await response.text())
            _logger.info(f"Scraped {len(res)} districts for city_id: {city_id}")

            if res:
                district_name_key = 'ArabicName' if use_arabic else 'EnglishName'
                return {str(city_id): [district[district_name_key] for district in res]}
    return {}


async def get_cities(regions: dict[str, str], use_arabic: bool = True) -> dict[str, list]:
    """
    Scrape all cities, which are hashes containing their arabic and English names
    as well as the Emirate region they belong in.
    :param regions:
    :param use_arabic:
    :return:
    """
    region_city_mapping = {}
    async with aiohttp.ClientSession() as session:
        async with session.post(
            'https://maps.splonline.com.sa/Home/GetCities',
            data=json.dumps({'cityId': 0}),
            headers=HEADERS
        ) as response:
            text = await response.text()
            if text:
                data = json.loads(text)
                _logger.info(f"Scraped {len(data)} cities")
                for city in data:
                    if city['fkEmirateID'] in regions:
                        region_name = regions[city['fkEmirateID']]
                        city_data = {
                            'name': city['ArabicName'] if use_arabic else city['EnglishName'],
                            'id': city['pkCityID'],
                            'districts': []
                        }
                        if region_name in region_city_mapping:
                            region_city_mapping[region_name].append(city_data)
                        else:
                            region_city_mapping[region_name] = [city_data]
                    else:
                        _logger.warning(f"Missing region for city:: {city}")
                return region_city_mapping


def save_to_excel(save_path: Path, json_source: Path):
    workbook = Workbook()
    header_style = NamedStyle(name='header_style')
    header_style.font = Font(bold=True, size=14)
    with open(json_source, "r", encoding='utf-8') as f:
        region_data_citi_final = json.loads(f.read())
        # Sort cities data by count of districts
        for idx, (region, cities) in enumerate(region_data_citi_final.items(), start=1):
            sheet = workbook.create_sheet(region, idx - 1)
            count = 1
            city_column_max_len = 1.2
            district_column_lens = {}
            sheet.column_dimensions[get_column_letter(1)].width = city_column_max_len
            sorted_cities = sorted(cities, key=lambda x: len(x.get('districts', [])), reverse=True)
            for city in sorted_cities:
                cell = sheet.cell(row=count, column=1, value=city['name'])
                cell.font = Font(bold=True, size=12)
                city_len = round(len(city['name']) * 1.2)
                if city_len > city_column_max_len:
                    city_column_max_len = city_len
                    sheet.column_dimensions[get_column_letter(1)].width = city_column_max_len

                for col_idx, district in enumerate(city['districts'], start=2):
                    sheet.cell(row=count, column=col_idx, value=district)
                    col_letter = get_column_letter(col_idx)
                    district_len = int(len(district) * 1)

                    col_max_len = district_column_lens.get(col_letter, 0) or district_len
                    if district_len > col_max_len or col_letter not in district_column_lens:
                        sheet.column_dimensions[col_letter].width = district_len
                        district_column_lens[col_letter] = district_len
                count += 1

    workbook.save(filename=save_path)


async def get_and_merge_districts(region_city_mapping: dict) -> dict:
    for region, cities in region_city_mapping.items():
        tasks = [get_districts(city['id']) for city in cities]
        get_districts_res = await asyncio.gather(*tasks)
        merged_districts = {
            city_id: dists for city_dist in get_districts_res for city_id, dists in city_dist.items() if city_dist
        }

        for city in cities:
            city['districts'] = merged_districts.get(city['id'], [])

    return region_city_mapping


async def main():
    regions = get_regions()
    _logger.info(f"Scraped {len(regions)} regions")
    region_city_mapping = await get_cities(regions)
    res_data = await get_and_merge_districts(region_city_mapping)

    parent = Path(__file__).resolve().parent

    json_results_path = parent / RESULTS_JSON
    with open(json_results_path, "w+", encoding='utf-8') as f:
        f.write(json.dumps(res_data))

    excel_results_path = parent / RESULTS_EXCEL
    save_to_excel(excel_results_path, json_results_path)


if __name__ == '__main__':
    asyncio.run(main())
